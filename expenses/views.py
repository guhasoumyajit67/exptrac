from django.views.generic.edit import CreateView, UpdateView, DeleteView, FormView
from django.views.generic import ListView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.db.models import ProtectedError, Count, Sum, Q
from django.shortcuts import redirect, render
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.views import View
from django.core.paginator import Paginator
from django.template.loader import render_to_string
import openpyxl
import re
import csv
import io
from datetime import datetime, timedelta

from .models import Transaction, Item, Payer, Category, StagingTransaction
from .forms import TransactionForm, ItemForm, PayerForm, ExcelUploadForm


# ==============================================================================
# HOMEPAGE VIEWS
# ==============================================================================

class HomePageView(TemplateView):
    template_name = "home.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        if not user.is_authenticated:
            return context

        now = timezone.now()
        current_year = now.year
        current_month = now.month

        tx_queryset = Transaction.objects.filter(
            user=user,
            date__year=current_year,
            date__month=current_month
        ).select_related('item', 'item__category', 'payer')

        total_outflow = tx_queryset.aggregate(total=Sum('price'))['total'] or 0.00
        context['total_outflow'] = total_outflow

        context['recent_transactions'] = tx_queryset.order_by('-date')

        category_data = tx_queryset.values('item__category__name').annotate(total=Sum('price')).order_by('-total')
        context['category_list_data'] = category_data

        if tx_queryset.exists():
            item_data = tx_queryset.values('item__name').annotate(total=Sum('price')).order_by('-total')[:6]
            context['category_labels'] = [entry['item__name'] for entry in item_data]
            context['category_amounts'] = [float(entry['total']) for entry in item_data]

            context['master_category_labels'] = [entry['item__category__name'] for entry in category_data[:6]]
            context['master_category_amounts'] = [float(entry['total']) for entry in category_data[:6]]

            if category_data:
                context['top_category_name'] = category_data[0]['item__category__name']
                context['top_category_amount'] = category_data[0]['total']

            frequency_map = tx_queryset.values('item__name').annotate(count=Count('id')).order_by('-count')
            if frequency_map:
                context['top_frequency_name'] = frequency_map[0]['item__name']
                context['top_frequency_count'] = frequency_map[0]['count']

            volume_map = tx_queryset.values('item__name', 'item__unit').annotate(volume=Sum('quantity')).order_by('-volume')
            if volume_map and volume_map[0]['volume']:
                context['top_volume_name'] = volume_map[0]['item__name']
                context['top_volume_count'] = volume_map[0]['volume']
                context['top_volume_unit'] = volume_map[0]['item__unit'] or "Units"
            else:
                context['top_volume_name'] = context.get('top_frequency_name', "None")
                context['top_volume_count'] = context.get('top_frequency_count', 0)
                context['top_volume_unit'] = "logs"

            trend_queryset = tx_queryset.values('date').annotate(total=Sum('price')).order_by('date')
            context['trend_labels'] = [entry['date'].strftime('%d %b') for entry in trend_queryset]
            
            running_sum = 0.0
            cumulative_amounts = []
            for entry in trend_queryset:
                running_sum += float(entry['total'])
                cumulative_amounts.append(running_sum)
                
            context['trend_amounts'] = cumulative_amounts

        else:
            context['category_labels'] = []
            context['category_amounts'] = []
            context['master_category_labels'] = []
            context['master_category_amounts'] = []
            context['trend_labels'] = []
            context['trend_amounts'] = []

        return context


# ==============================================================================
# TRANSACTION VIEWS
# ==============================================================================

class TransactionCreateView(LoginRequiredMixin, CreateView):
    model = Transaction
    form_class = TransactionForm
    template_name = "expenses/transaction_form.html"
    success_url = reverse_lazy("create_transaction")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs
    
    def get_initial(self):
        initial = super().get_initial()
        initial_date = timezone.now().date()
        initial_payer = None

        last_transaction = (
            Transaction.objects.filter(user=self.request.user)
            .order_by("-id")
            .first()
        )
        if last_transaction:
            initial_date = last_transaction.date
            initial_payer = last_transaction.payer

        initial["date"] = initial_date
        if initial_payer:
            initial["payer"] = initial_payer

        return initial
    
    def form_valid(self, form):
        form.instance.user = self.request.user
        messages.success(
            self.request, 
            f"Successfully logged '{form.instance.item.name}' for ₹{form.instance.price}!"
        )
        return super().form_valid(form)


class TransactionUpdateView(LoginRequiredMixin, UpdateView):
    model = Transaction
    form_class = TransactionForm
    template_name = "expenses/transaction_form.html"
    success_url = reverse_lazy("transaction_list")

    def get_queryset(self):
        return Transaction.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs


class TransactionDeleteView(LoginRequiredMixin, DeleteView):
    model = Transaction
    template_name = "expenses/transaction_confirm_delete.html"
    success_url = reverse_lazy("transaction_list")

    def get_queryset(self):
        return Transaction.objects.filter(user=self.request.user)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        item_name = self.object.item.name
        success_url = self.get_success_url()
        self.object.delete()
        messages.success(request, f"Transaction '{item_name}' has been successfully deleted.")
        return redirect(success_url)


class TransactionListView(LoginRequiredMixin, ListView):
    model = Transaction
    template_name = "expenses/transaction_list.html"
    context_object_name = "transactions"
    paginate_by = 50

    def get_queryset(self):
        queryset = (
            Transaction.objects.filter(user=self.request.user)
            .select_related("item__category", "payer")
            .order_by("-date", "-id")
        )
        
        # Get filter parameters from request
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        period = self.request.GET.get('period')
        search = self.request.GET.get('search')
        category = self.request.GET.get('category')
        payer = self.request.GET.get('payer')
        
        # Apply date range filter
        if date_from:
            try:
                queryset = queryset.filter(date__gte=date_from)
            except ValueError:
                pass
        if date_to:
            try:
                queryset = queryset.filter(date__lte=date_to)
            except ValueError:
                pass
        
        # Apply quick period filter
        if period and period != 'all':
            today = timezone.now().date()
            if period == '1w':
                start_date = today - timedelta(days=7)
            elif period == '1m':
                start_date = today - timedelta(days=30)
            elif period == '3m':
                start_date = today - timedelta(days=90)
            elif period == '1y':
                start_date = today - timedelta(days=365)
            else:
                start_date = None
            
            if start_date:
                queryset = queryset.filter(date__gte=start_date)
        
        # Apply search filter
        if search:
            queryset = queryset.filter(
                Q(item__name__icontains=search) |
                Q(payer__name__icontains=search) |
                Q(item__category__name__icontains=search) |
                Q(comment__icontains=search)
            )
        
        # Apply category filter
        if category:
            queryset = queryset.filter(item__category__name=category)
        
        # Apply payer filter
        if payer:
            queryset = queryset.filter(payer__name=payer)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass filter params back to template for persistence
        context['filters'] = {
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', ''),
            'period': self.request.GET.get('period', 'all'),
            'search': self.request.GET.get('search', ''),
            'category': self.request.GET.get('category', ''),
            'payer': self.request.GET.get('payer', ''),
        }
        return context


# ==============================================================================
# EXPORT VIEW
# ==============================================================================

class TransactionExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        format_type = request.GET.get('format', 'csv')
        
        # Get all transactions for the user with filters
        queryset = Transaction.objects.filter(
            user=request.user
        ).select_related('item', 'item__category', 'payer').order_by('-date')
        
        # Apply filters from request (same as TransactionListView)
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        period = request.GET.get('period')
        search = request.GET.get('search')
        category = request.GET.get('category')
        payer = request.GET.get('payer')
        
        # Apply date range filter
        if date_from:
            try:
                queryset = queryset.filter(date__gte=date_from)
            except ValueError:
                pass
        if date_to:
            try:
                queryset = queryset.filter(date__lte=date_to)
            except ValueError:
                pass
        
        # Apply quick period filter
        if period and period != 'all':
            today = timezone.now().date()
            if period == '1w':
                start_date = today - timedelta(days=7)
            elif period == '1m':
                start_date = today - timedelta(days=30)
            elif period == '3m':
                start_date = today - timedelta(days=90)
            elif period == '1y':
                start_date = today - timedelta(days=365)
            else:
                start_date = None
            
            if start_date:
                queryset = queryset.filter(date__gte=start_date)
        
        # Apply search filter
        if search:
            queryset = queryset.filter(
                Q(item__name__icontains=search) |
                Q(payer__name__icontains=search) |
                Q(item__category__name__icontains=search) |
                Q(comment__icontains=search)
            )
        
        # Apply category filter
        if category:
            queryset = queryset.filter(item__category__name=category)
        
        # Apply payer filter
        if payer:
            queryset = queryset.filter(payer__name=payer)
        
        # Prepare data
        data = []
        for t in queryset:
            data.append({
                'Date': t.date.strftime('%Y-%m-%d'),
                'Category': t.item.category.name if t.item.category else '',
                'Item': t.item.name,
                'Price': float(t.price),
                'Quantity': float(t.quantity) if t.quantity else '',
                'Unit': t.item.unit or '',
                'Paid_By': t.payer.name,
                'Comment': t.comment or ''
            })
        
        # Export based on format
        if format_type == 'csv':
            return self.export_csv(data)
        elif format_type == 'excel':
            return self.export_excel(data)
        elif format_type == 'pdf':
            return self.export_pdf(data)
        else:
            return JsonResponse({'error': 'Unsupported format'}, status=400)
    
    def export_csv(self, data):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="transactions_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        if data:
            # Write headers
            headers = list(data[0].keys())
            writer.writerow(headers)
            # Write data
            for row in data:
                writer.writerow(row.values())
        else:
            writer.writerow(['No transactions found'])
        
        return response
    
    def export_excel(self, data):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return JsonResponse({'error': 'openpyxl not installed'}, status=500)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        if data:
            headers = list(data[0].keys())
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = thin_border
            
            # Auto-width columns
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add total row count
            total_row = len(data) + 2
            ws.cell(row=total_row, column=1, value=f"Total: {len(data)} transactions")
            ws.cell(row=total_row, column=1).font = Font(bold=True)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="transactions_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    
    def export_pdf(self, data):
        # Generate HTML content for PDF
        context = {
            'data': data,
            'total': len(data),
            'generated_date': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
        html_string = render_to_string('expenses/export_pdf.html', context)
        
        try:
            from weasyprint import HTML, CSS
            from weasyprint.text.fonts import FontConfiguration
            
            font_config = FontConfiguration()
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="transactions_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            
            HTML(string=html_string).write_pdf(
                response,
                font_config=font_config,
                stylesheets=[CSS(string='@page { size: A4 landscape; margin: 1cm; }')]
            )
            return response
        except ImportError:
            # Fallback: Use reportlab if weasyprint not available
            return self.export_pdf_reportlab(data)
    
    def export_pdf_reportlab(self, data):
        """Fallback PDF generation using reportlab"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
        except ImportError:
            return JsonResponse({'error': 'reportlab not installed'}, status=500)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="transactions_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30,
        )
        
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#0D6EFD'),
            spaceAfter=12
        )
        elements.append(Paragraph('Transaction Report', title_style))
        
        # Info
        info_style = ParagraphStyle(
            'Info',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=8
        )
        elements.append(Paragraph(f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}', info_style))
        elements.append(Paragraph(f'Total Transactions: {len(data)}', info_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Table data
        if data:
            headers = list(data[0].keys())
            table_data = [headers]
            
            for row in data:
                table_data.append([row.get(h, '') for h in headers])
            
            # Create table
            table = Table(table_data, repeatRows=1)
            
            # Style table
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ])
            
            # Set column widths
            col_widths = []
            for i in range(len(headers)):
                col_widths.append(1.0 * inch)
            table._argW = col_widths
            
            table.setStyle(table_style)
            elements.append(table)
            
            # Total
            elements.append(Spacer(1, 0.1 * inch))
            total_style = ParagraphStyle(
                'Total',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#0D6EFD'),
                alignment=2  # Right align
            )
            elements.append(Paragraph(f'Total: {len(data)} transactions', total_style))
        else:
            elements.append(Paragraph('No transactions found', styles['Normal']))
        
        # Footer
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1  # Center align
        )
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph('Generated by ExpTrac - Expense Tracker', footer_style))
        
        # Build PDF
        doc.build(elements)
        
        # Get value from buffer
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response


# ==============================================================================
# ITEM VIEWS
# ==============================================================================

class ItemManagementListView(LoginRequiredMixin, ListView):
    model = Item
    template_name = "expenses/manage_items.html"
    context_object_name = "items"

    def get_queryset(self):
        return Item.objects.filter(user=self.request.user)\
                           .select_related('category')\
                           .annotate(transaction_count=Count('transaction'))\
                           .order_by('name')


class ItemCreateView(LoginRequiredMixin, CreateView):
    model = Item
    form_class = ItemForm
    template_name = "expenses/item_form.html"

    def get_success_url(self):
        redirect_to = self.request.GET.get('next') or self.request.POST.get('next')
        if redirect_to:
            return redirect_to
        return reverse_lazy("create_transaction")

    def form_valid(self, form):
        item_name = form.cleaned_data.get('name')
        if Item.objects.filter(user=self.request.user, name__iexact=item_name).exists():
            form.add_error('name', f'You have already created a custom item named "{item_name}".')
            return self.form_invalid(form)
            
        form.instance.user = self.request.user
        return super().form_valid(form)


class ItemUpdateView(LoginRequiredMixin, UpdateView):
    model = Item
    form_class = ItemForm
    template_name = "expenses/item_form.html"

    def get_queryset(self):
        return Item.objects.filter(user=self.request.user)

    def get_success_url(self):
        redirect_to = self.request.GET.get('next') or self.request.POST.get('next')
        if redirect_to:
            return redirect_to
        return reverse_lazy("manage_items")
    

class ItemDeleteView(LoginRequiredMixin, DeleteView):
    model = Item
    template_name = "expenses/item_confirm_delete.html"

    def get_queryset(self):
        return Item.objects.filter(user=self.request.user)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        item_name = self.object.name
        redirect_to = request.POST.get('next')
        fallback_url = redirect_to if redirect_to else reverse("manage_items")
        
        try:
            self.object.delete()
            messages.success(request, f"Success! Item '{item_name}' has been safely removed.")
            return redirect(fallback_url)
        except ProtectedError:
            messages.error(
                request, 
                f"Cannot delete item '{item_name}' because it is currently tracked inside active ledger logs!"
            )
            return redirect(fallback_url)


# ==============================================================================
# PAYER VIEWS
# ==============================================================================

class PayerManagementListView(LoginRequiredMixin, ListView):
    model = Payer
    template_name = "expenses/manage_payers.html"
    context_object_name = "payers"

    def get_queryset(self):
        return (
            Payer.objects.filter(user=self.request.user)
            .annotate(transaction_count=Count('transaction'))
            .order_by('name')
        )


class PayerCreateView(LoginRequiredMixin, CreateView):
    model = Payer
    form_class = PayerForm
    template_name = "expenses/payer_form.html"

    def get(self, request, *args, **kwargs):
        referer = request.META.get('HTTP_REFERER')
        if referer:
            request.session['payer_form_origin_url'] = referer
        return super().get(request, *args, **kwargs)

    def get_success_url(self):
        return self.request.session.pop('payer_form_origin_url', reverse("manage_payers"))

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class PayerUpdateView(LoginRequiredMixin, UpdateView):
    model = Payer
    form_class = PayerForm
    template_name = "expenses/payer_form.html"

    def get_queryset(self):
        return Payer.objects.filter(user=self.request.user)

    def get_success_url(self):
        redirect_to = self.request.GET.get('next') or self.request.POST.get('next')
        if redirect_to:
            return redirect_to
        return reverse_lazy("manage_payers")


class PayerDeleteView(LoginRequiredMixin, DeleteView):
    model = Payer
    template_name = "expenses/payer_confirm_delete.html"

    def get_queryset(self):
        return Payer.objects.filter(user=self.request.user)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        payer_name = self.object.name
        
        redirect_to = request.POST.get('next')
        fallback_url = redirect_to if redirect_to else reverse("manage_payers")
        
        try:
            self.object.delete()
            messages.success(request, f"Success! Payer '{payer_name}' has been safely removed.")
            return redirect(fallback_url)
            
        except ProtectedError:
            messages.error(
                request, 
                f"Cannot delete payer '{payer_name}' because they are currently tied to existing transactions in your ledger."
            )
            return redirect(fallback_url)


# ==============================================================================
# BULK OPERATIONS VIEWS
# ==============================================================================

class TransactionBulkUploadView(LoginRequiredMixin, FormView):
    form_class = ExcelUploadForm
    template_name = "expenses/bulk_upload.html"
    success_url = reverse_lazy("bulk_upload_review")

    def form_valid(self, form):
        excel_file = self.request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
        
        item_cache = {item.name.strip().lower(): item for item in Item.objects.all()}
        payer_cache = {payer.name.strip().lower(): payer for payer in Payer.objects.filter(user=self.request.user)}
        
        StagingTransaction.objects.filter(user=self.request.user).delete()
        
        staging_pool = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[0] is None:
                continue
                
            date_val, item_str, price_raw, quantity_raw, payer_str, comment_val = row[:6]
            item_raw_name = str(item_str).strip() if item_str else ""
            payer_raw_name = str(payer_str).strip() if payer_str else ""

            final_price = 0.0
            if price_raw is not None:
                price_digits = re.sub(r'[^\d.]', '', str(price_raw))
                final_price = float(price_digits) if price_digits else 0.0

            final_quantity = None
            if quantity_raw is not None and str(quantity_raw).strip():
                qty_match = re.search(r'\d+(\.\d+)?', str(quantity_raw))
                if qty_match:
                    qty_float = float(qty_match.group())
                    final_quantity = int(qty_float) if qty_float.is_integer() else qty_float

            item_obj = item_cache.get(item_raw_name.lower())
            payer_obj = payer_cache.get(payer_raw_name.lower())
            date_str = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val)

            errors = []
            if not item_obj:
                errors.append("Missing Item")
            if not payer_obj:
                errors.append("Missing Payer")

            staging_pool.append(
                StagingTransaction(
                    user=self.request.user,
                    row_idx=row_idx,
                    date=date_str,
                    item_name=item_raw_name,
                    item_id=item_obj.id if item_obj else None,
                    price=final_price,
                    quantity=final_quantity,
                    payer_name=payer_raw_name,
                    payer_id=payer_obj.id if payer_obj else None,
                    comment=str(comment_val).strip() if comment_val else "",
                    error=" & ".join(errors)
                )
            )

        if staging_pool:
            StagingTransaction.objects.bulk_create(staging_pool)

        return redirect('bulk_upload_review')


class TransactionBulkReviewView(LoginRequiredMixin, View):
    template_name = "expenses/bulk_review.html"

    def get(self, request, *args, **kwargs):
        user_rows = StagingTransaction.objects.filter(user=request.user)
        review_rows = user_rows.exclude(error="")
        valid_rows_count = user_rows.filter(error="").count()
        
        if not user_rows.exists():
            messages.info(request, "No transaction staging queue found. Please upload a spreadsheet first.")
            return redirect('bulk_upload_transactions')

        missing_items = []
        missing_payers = []
        has_blank_payers = False  

        for row in review_rows:
            if row.item_name and not row.item_id:
                cleaned_item = row.item_name.strip()
                if cleaned_item not in missing_items:
                    missing_items.append(cleaned_item)
            
            if not row.payer_id:
                if not row.payer_name or not str(row.payer_name).strip():
                    has_blank_payers = True
                else:
                    cleaned_payer = row.payer_name.strip()
                    if cleaned_payer not in missing_payers:
                        missing_payers.append(cleaned_payer)
                    
        context = {
            'missing_items': sorted(missing_items),
            'missing_payers': sorted(missing_payers),
            'has_blank_payers': has_blank_payers,  
            'review_rows_count': review_rows.count(),
            'valid_rows_count': valid_rows_count,
            'all_items': Item.objects.all().order_by('name'),
            'all_payers': Payer.objects.filter(user=request.user).order_by('name'),
            'all_categories': Category.objects.all().order_by('name'),
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        target_item_name = request.POST.get('target_item_name', '').strip()
        target_payer_name = request.POST.get('target_payer_name', '').strip()

        if action == "resolve_blank_payers":
            existing_payer_id = request.POST.get('existing_payer_id')
            existing_payer = Payer.objects.get(id=existing_payer_id, user=request.user)
            
            blank_rows = StagingTransaction.objects.filter(user=request.user, payer_id__isnull=True)
            for row in blank_rows:
                if not row.payer_name or not str(row.payer_name).strip():
                    row.payer_id = existing_payer.id
                    row.payer_name = existing_payer.name
                    row.error = "Missing Item" if not row.item_id else ""
                    row.save()
            
            messages.success(request, f"Assigned payer profile '{existing_payer.name}' to all unassigned rows.")
            return redirect('bulk_upload_review')

        elif action == "create_custom":
            category_obj = Category.objects.get(id=request.POST.get('category_id'))
            new_item = Item.objects.create(name=target_item_name, user=request.user, category=category_obj, unit=request.POST.get('unit'))
            
            matches = StagingTransaction.objects.filter(user=request.user, item_name__iexact=target_item_name)
            for row in matches:
                row.item_id = new_item.id
                row.error = "Missing Payer" if not row.payer_id else ""
                row.save()
                
            messages.success(request, f"Created Custom Item '{target_item_name}'.")
            return redirect('bulk_upload_review')

        elif action == "map_existing":
            existing_item = Item.objects.get(id=request.POST.get('existing_item_id'))
            
            matches = StagingTransaction.objects.filter(user=request.user, item_name__iexact=target_item_name)
            for row in matches:
                row.item_id = existing_item.id
                row.item_name = existing_item.name
                row.error = "Missing Payer" if not row.payer_id else ""
                row.save()
                
            messages.success(request, f"Mapped lines to tracking item '{existing_item.name}'.")
            return redirect('bulk_upload_review')

        elif action == "create_custom_payer":
            new_payer = Payer.objects.create(name=target_payer_name, user=request.user)
            
            matches = StagingTransaction.objects.filter(user=request.user, payer_name__iexact=target_payer_name)
            for row in matches:
                row.payer_id = new_payer.id
                row.error = "Missing Item" if not row.item_id else ""
                row.save()
                
            messages.success(request, f"Created Payer profile '{target_payer_name}'.")
            return redirect('bulk_upload_review')

        elif action == "map_existing_payer":
            existing_payer = Payer.objects.get(id=request.POST.get('existing_payer_id'), user=request.user)
            
            matches = StagingTransaction.objects.filter(user=request.user, payer_name__iexact=target_payer_name)
            for row in matches:
                row.payer_id = existing_payer.id
                row.payer_name = existing_payer.name
                row.error = "Missing Item" if not row.item_id else ""
                row.save()
                
            messages.success(request, f"Mapped lines to payer '{existing_payer.name}'.")
            return redirect('bulk_upload_review')

        elif action == "skip_item":
            StagingTransaction.objects.filter(user=request.user, item_name__iexact=target_item_name).delete()
            messages.info(request, f"Discarded spreadsheet lines containing '{target_item_name}'.")
            return redirect('bulk_upload_review')

        if StagingTransaction.objects.filter(user=request.user).exclude(error="").exists():
            return redirect('bulk_upload_review')

        return redirect('bulk_upload_confirm_commit')


class BulkUploadCommitView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        valid_staging_rows = StagingTransaction.objects.filter(user=request.user, error="")
        
        if not valid_staging_rows.exists():
            messages.info(request, "No pending valid records to save.")
            return redirect('transaction_list')

        transactions_to_create = [
            Transaction(
                user=request.user,
                date=row.date,
                item_id=row.item_id,
                price=row.price,
                quantity=row.quantity,
                payer_id=row.payer_id,
                comment=row.comment
            ) for row in valid_staging_rows
        ]

        if transactions_to_create:
            Transaction.objects.bulk_create(transactions_to_create)
            messages.success(request, f"Success! Safely committed {len(transactions_to_create)} rows into your ledger.")

        StagingTransaction.objects.filter(user=request.user).delete()
        return redirect('transaction_list')


class TransactionBulkDeleteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        transaction_ids = request.POST.getlist('transaction_ids')
        if not transaction_ids:
            messages.warning(request, "No transactions were selected for deletion.")
            return redirect('transaction_list')
            
        target_queryset = Transaction.objects.filter(id__in=transaction_ids, user=request.user)
        total_selected = target_queryset.count()
        deleted_count = 0
        
        if total_selected > 0:
            for transaction in target_queryset:
                try:
                    transaction.delete()
                    deleted_count += 1
                except Exception:
                    pass

        if deleted_count > 0:
            messages.success(request, f"Successfully deleted {deleted_count} selected transactions.")
            if deleted_count < total_selected:
                ignored = total_selected - deleted_count
                messages.warning(request, f"{ignored} transactions were protected by database constraints.")
        else:
            messages.error(request, "Failed to delete transactions. Selected records are protected or could not be found.")
            
        return redirect('transaction_list')


class TransactionDeleteStatusView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        progress = request.session.get('delete_progress', 0)
        return JsonResponse({'progress': progress})